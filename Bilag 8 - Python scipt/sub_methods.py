import openpyxl
import pandas as pd


def purge_data(file_template):
    book = openpyxl.load_workbook(file_template)
    sheet = book['Template']
    sheet.delete_rows(2, sheet.max_row)
    book.save(file_template)
    return sheet


def normalize_data(filename, sheet_index):
    v1_data_set = pd.read_excel(filename, sheet_name=sheet_index, engine="openpyxl")
    df = pd.DataFrame(v1_data_set)

    if(pd.isna(df.iloc[0,0])):
        df = clean_empty_rows(df)

    if determine_header_direction(df):
        df = df.transpose()
        df.reset_index(level=0, inplace=True)
        df.columns = df.iloc[0]
        df = df[1:]

    return df


def clean_empty_rows(df):
    df.dropna(axis=1, how='all', inplace=True)
    df.dropna(axis=0, how='all', inplace=True)
    df.reset_index(drop=True, inplace=True)
    df.columns = df.iloc[0]
    df.drop(df.index[0], inplace=True)
    return df


def determine_header_direction(df):
    header_list = ["Number", "Rumnavne", "Area", "Specified Supply Airflow", "Specified Return Airflow", "Luftmængde",
                   "Room: Department", "Min. Invendig Diamenter (Tryktabsgradient 0,7)", "Indendig Diameter"]
    
    for header in header_list:
        if header == df.iat[1,0]:
            return True
    return False


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def get_column_index(name, sheet):
    for column_name in sheet.iter_cols(1, sheet.max_column):
        if column_name[0].value == name:
            return column_name[0].col_idx - 1


def append_to_excel(start_column, column_name, file_template, df):
    append_df_to_excel(file_template, df, startrow=1, startcol=start_column, sheet_name='Template',
                       truncate_sheet=False, header=False, columns=[column_name], index=False)